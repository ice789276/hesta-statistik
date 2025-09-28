import os
import io
import pandas as pd
import numpy as np
import altair as alt
import streamlit as st
from datetime import datetime
from dotenv import load_dotenv
import openpyxl
from openpyxl.styles import PatternFill

# .env laden
load_dotenv()
PASS = st.secrets("ADMIN_PASS")

DATA_FILE = "data.csv"

LAENDER_INPUT = [
    "🇨🇭 Schweiz", "🇩🇪 Deutschland", "🇫🇷 Frankreich", "🇮🇹 Italien",
    "🇷🇺 Russland", "🇪🇸 Spanien", "🇬🇧 Grossbritannien", "🇺🇸 USA",
    "🇨🇦 Kanada", "🇧🇷 Brasilien", "🇨🇳 China", "🇭🇰 Hongkong",
    "🇯🇵 Japan", "🇰🇷 Korea", "🇹🇭 Thailand", "🇦🇪 VAE", "🇦🇺 Australien"
]
LAENDER_REIHENFOLGE = LAENDER_INPUT + ["Total ausgewählte Länder", "Total übrige Länder", "Total"]

if "auth" not in st.session_state:
    st.session_state.auth = False
if "login_attempted" not in st.session_state:
    st.session_state.login_attempted = False

st.title("Hesta-Statistik Hotel Savoy Bern")

if not st.session_state.auth:
    pw_input = st.text_input("Passwort eingeben", type="password")
    if st.button("Login"):
        st.session_state.login_attempted = True
        if pw_input == PASS:
            st.session_state.auth = True
        else:
            st.error("Falsches Passwort")

    # Stoppt alles andere, wenn noch nicht eingeloggt
    if not st.session_state.auth:
        st.stop()

# Ab hier: nur sichtbar nach Login
st.success("Login erfolgreich ✅")


# Ab hier nur nach Login
st.title("Hesta-Statistik Hotel Savoy Bern")

# =========================
# Datei & DataFrame laden
# =========================
if os.path.exists(DATA_FILE):
    try:
        df = pd.read_csv(DATA_FILE, parse_dates=["Monat"])
    except pd.errors.EmptyDataError:
        df = pd.DataFrame(columns=["Monat", "Land", "Anreisen", "Nächte"])
else:
    df = pd.DataFrame(columns=["Monat", "Land", "Anreisen", "Nächte"])

# =========================
# Eingabeformular
# =========================
with st.form("eingabe_formular"):
    jahr = st.selectbox(
        "Jahr",
        list(range(2017, datetime.today().year + 2)),
        index=list(range(2017, datetime.today().year + 2)).index(datetime.today().year)
    )
    monat = st.selectbox("Monat", list(range(1, 13)), index=datetime.today().month - 2)

    anreisen_inputs = {}
    naechte_inputs = {}

    # Labels über den Eingabefeldern
    col0, col1, col2 = st.columns([2, 1, 1])
    with col0: st.markdown("<b>Land</b>", unsafe_allow_html=True)
    with col1: st.markdown("<b>Anreisen</b>", unsafe_allow_html=True)
    with col2: st.markdown("<b>Nächte</b>", unsafe_allow_html=True)

    # Länder-Eingaben mit abwechselnder Hintergrundfarbe
    for i, land in enumerate(LAENDER_INPUT):
        bg_color = "#f2f2f2" if i % 2 == 0 else "white"
        col0, col1, col2 = st.columns([2, 1, 1])
        with col0:
            st.markdown(f"<div style='background-color:{bg_color}; font-size:20px; height:38px; display:flex; align-items:flex-end;'>{land}</div>", unsafe_allow_html=True)
        with col1:
            val_anreisen = st.number_input("", min_value=0, value=0, step=1, key=f"{land}_anreisen", label_visibility="collapsed")
            anreisen_inputs[land] = "" if val_anreisen == 0 else val_anreisen
        with col2:
            val_naechte = st.number_input("", min_value=0, value=0, step=1, key=f"{land}_naechte", label_visibility="collapsed")
            naechte_inputs[land] = "" if val_naechte == 0 else val_naechte

        # Hintergrundfarbe der Eingabefelder
        st.markdown(f"""
            <style>
            div[data-testid="stNumberInput"] {{
                background-color: {bg_color} !important;
            }}
            </style>
        """, unsafe_allow_html=True)

    # Berechnung Totals
    total_ausgewählte_anreisen = sum(int(anreisen_inputs[l]) if anreisen_inputs[l] != "" else 0 for l in LAENDER_INPUT)
    total_ausgewählte_naechte = sum(int(naechte_inputs[l]) if naechte_inputs[l] != "" else 0 for l in LAENDER_INPUT)

    col1, col2, col3 = st.columns([2, 1, 1])
    with col1: st.markdown("<div style='font-size:20px; background-color:#d9edf7;'>Total ausgewählte Länder</div>", unsafe_allow_html=True)
    with col2: st.number_input("", value=total_ausgewählte_anreisen, disabled=True, key="total_ausgewählte_anreisen")
    with col3: st.number_input("", value=total_ausgewählte_naechte, disabled=True, key="total_ausgewählte_naechte")

    if "total_anreisen_input" not in st.session_state:
        st.session_state.total_anreisen_input = total_ausgewählte_anreisen
    if "total_naechte_input" not in st.session_state:
        st.session_state.total_naechte_input = total_ausgewählte_naechte

    col1, col2, col3 = st.columns([2, 1, 1])
    with col1: st.markdown("<div style='font-size:20px; background-color:#d9edf7;'>Total</div>", unsafe_allow_html=True)
    with col2: total_anreisen_input = st.number_input("", min_value=0, step=1, key="total_anreisen_input")
    with col3: total_naechte_input = st.number_input("", min_value=0, step=1, key="total_naechte_input")

    total_übrige_anreisen = total_anreisen_input - total_ausgewählte_anreisen
    total_übrige_naechte = total_naechte_input - total_ausgewählte_naechte
    col1, col2, col3 = st.columns([2, 1, 1])
    with col1: st.markdown("<div style='font-size:20px; background-color:#d9edf7;'>Total übrige Länder</div>", unsafe_allow_html=True)
    with col2: st.number_input("", value=total_übrige_anreisen, disabled=True, key="total_übrige_anreisen")
    with col3: st.number_input("", value=total_übrige_naechte, disabled=True, key="total_übrige_naechte")

    submitted = st.form_submit_button("Speichern")

# =========================
# Daten speichern
# =========================
if submitted:
    monat_datum = pd.Timestamp(datetime(jahr, monat, 1))
    daten_neu = []
    for land in LAENDER_INPUT:
        daten_neu.append({"Monat": monat_datum, "Land": land,
                          "Anreisen": int(anreisen_inputs[land] or 0),
                          "Nächte": int(naechte_inputs[land] or 0)})
    daten_neu += [
        {"Monat": monat_datum, "Land": "Total ausgewählte Länder", "Anreisen": total_ausgewählte_anreisen, "Nächte": total_ausgewählte_naechte},
        {"Monat": monat_datum, "Land": "Total übrige Länder", "Anreisen": total_übrige_anreisen, "Nächte": total_übrige_naechte},
        {"Monat": monat_datum, "Land": "Total", "Anreisen": total_anreisen_input, "Nächte": total_naechte_input}
    ]
    df_neu = pd.DataFrame(daten_neu)
    df = pd.concat([df[~df.set_index(['Monat','Land']).index.isin(df_neu.set_index(['Monat','Land']).index)], df_neu], ignore_index=True)
    df["Land"] = pd.Categorical(df["Land"], categories=LAENDER_REIHENFOLGE, ordered=True)
    df = df.sort_values(by=["Monat","Land"])
    df.to_csv(DATA_FILE, index=False)
    st.success("Daten erfolgreich gespeichert ✅")

# =========================
# Tabellen & Diagramme
# =========================
if not df.empty:
    # Pivot-Tabellen
    pivot_naechte = df.pivot_table(index="Monat", columns="Land", values="Nächte", aggfunc="sum", fill_value=0)
    pivot_anreisen = df.pivot_table(index="Monat", columns="Land", values="Anreisen", aggfunc="sum", fill_value=0)

    # Durchschnittliche Nächte pro Anreise
    avg_naechte_pro_anreise = pivot_naechte / pivot_anreisen.replace(0, np.nan)

    # Prozentualer Anteil der Nächte
    total_naechte = pivot_naechte.sum(axis=1)
    percent_naechte = pivot_naechte.div(total_naechte, axis=0) * 100

    # Vorjahresmonat: exakt derselbe Monat des Vorjahres
    def prev_year_month(df, month_date):
        try:
            return df.loc[month_date - pd.DateOffset(years=1)]
        except KeyError:
            # Wenn kein Vorjahresmonat existiert, alle Werte 0
            return pd.Series(0, index=df.columns)

    # Berechnung % Veränderung zum Vorjahr
    mom_change_year = pd.DataFrame(index=pivot_naechte.index, columns=pivot_naechte.columns)
    for current_month in pivot_naechte.index:
        prev_values = prev_year_month(pivot_naechte, current_month)
        current_values = pivot_naechte.loc[current_month]
        mom_change_year.loc[current_month] = ((current_values - prev_values) / prev_values.replace(0, np.nan)) * 100

    # YTD-Berechnungen
    ytd_naechte = pivot_naechte.cumsum()
    ytd_anreisen = pivot_anreisen.cumsum()
    ytd_avg = ytd_naechte / ytd_anreisen.replace(0, np.nan)
    ytd_percent = ytd_naechte.div(ytd_naechte.sum(axis=1), axis=0) * 100

    # YTD % Veränderung zum Vorjahr (gleicher Monat)
    ytd_change_year = pd.DataFrame(index=pivot_naechte.index, columns=pivot_naechte.columns)
    for current_month in ytd_naechte.index:
        prev_values = prev_year_month(ytd_naechte, current_month)
        current_values = ytd_naechte.loc[current_month]
        ytd_change_year.loc[current_month] = ((current_values - prev_values) / prev_values.replace(0, np.nan)) * 100


    # Monat auswählen, Standard = letzter eingetragener Monat
    letzter_monat = df['Monat'].max()  # letzter Monat im DataFrame
    monat_auswahl = st.selectbox(
        "Monat auswählen",
        [d.strftime('%Y-%m') for d in pivot_naechte.index],
        index=list(pivot_naechte.index).index(letzter_monat)  # setzt den Index auf den letzten Monat
    )
    monat_date = pd.to_datetime(monat_auswahl)


    # Anzeige DataFrame vorbereiten
    temp = pd.DataFrame({
        'Land': pivot_naechte.columns,
        'Durchschnittliche Nächte pro Anreise': avg_naechte_pro_anreise.loc[monat_date].fillna(0).values,
        '% Anteil Nächte': percent_naechte.loc[monat_date].values,
        '% Veränderung zum Vorjahr': mom_change_year.loc[monat_date].values,
        'YTD Durchschnittliche Nächte pro Anreise': ytd_avg.loc[monat_date].fillna(0).values,
        'YTD % Anteil Nächte': ytd_percent.loc[monat_date].values,
        'YTD % Veränderung zum Vorjahr': ytd_change_year.loc[monat_date].values
    })

    # Reihenfolge fixieren
    temp = temp.set_index('Land').reindex(LAENDER_REIHENFOLGE).reset_index()

    # Styling-Funktion
    def highlight_change(val):
        if pd.isna(val):
            return ''
        elif val > 0:
            return 'background-color: #ccffcc'
        elif val < 0:
            return 'background-color: #ffcccc'
        return ''

    # Numerische Spalten formatieren
    numeric_cols = [
        'Durchschnittliche Nächte pro Anreise',
        '% Anteil Nächte',
        '% Veränderung zum Vorjahr',
        'YTD Durchschnittliche Nächte pro Anreise',
        'YTD % Anteil Nächte',
        'YTD % Veränderung zum Vorjahr'
    ]

    # Auf 2 Nachkommastellen runden und NaN als "-" darstellen
    for col in numeric_cols:
        temp[col] = temp[col].apply(lambda x: f"{x:.2f}" if pd.notnull(x) else "-")

    # Styling-Funktion für Farben
    def highlight_change(val):
        try:
            val_float = float(val)
        except:
            return ""
        if val_float > 0:
            return 'background-color: #ccffcc'  # hellgrün
        elif val_float < 0:
            return 'background-color: #ffcccc'  # hellrot
        return ''

    styled_df = temp.style.applymap(
        highlight_change,
        subset=['% Veränderung zum Vorjahr', 'YTD % Veränderung zum Vorjahr']
    )

    st.header(f"Berechnete Kennzahlen für {monat_auswahl}")
    st.dataframe(styled_df, height=750, width=1200, use_container_width=False)



    # Excel-Datei in BytesIO vorbereiten
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        temp.to_excel(writer, index=False, sheet_name='Monatliche Daten')
        workbook = writer.book
        sheet = writer.sheets['Monatliche Daten']

        # Farben definieren
        gruen = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
        rot = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")

        # Spaltennamen der %-Veränderungen
        perc_cols = ['% Veränderung zum Vorjahr', 'YTD % Veränderung zum Vorjahr']
        col_idx_map = {cell.value: cell.column for cell in sheet[1] if cell.value in perc_cols}

        # Werte farbig hinterlegen
        for row_idx in range(2, sheet.max_row + 1):
            for col_name, col_idx in col_idx_map.items():
                value = sheet.cell(row=row_idx, column=col_idx).value
                try:
                    num = float(value)
                    if num > 0:
                        sheet.cell(row=row_idx, column=col_idx).fill = gruen
                    elif num < 0:
                        sheet.cell(row=row_idx, column=col_idx).fill = rot
                except (ValueError, TypeError):
                    continue  # Wenn kein numerischer Wert, ignorieren

    st.download_button(
        label="Tabelle als XLSX herunterladen",
        data=output.getvalue(),
        file_name=f'Tabelle_{monat_auswahl}_farben.xlsx',
        mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

    # Balkendiagramm pro Land
    land_auswahl = st.selectbox("Land für Balkendiagramm auswählen", pivot_naechte.columns)
    df_chart = pivot_naechte[[land_auswahl]].reset_index().rename(columns={land_auswahl:'Nächte'})
    df_chart['Monat'] = df_chart['Monat'].dt.strftime('%B %Y')
    df_chart = df_chart.sort_values('Monat')
    chart = alt.Chart(df_chart).mark_bar().encode(
        x=alt.X('Monat:N', sort=list(df_chart['Monat'])),
        y=alt.Y('Nächte:Q'),
        tooltip=['Monat','Nächte']
    ).properties(width=700)
    st.altair_chart(chart)

    # YTD-Balkendiagramm
    df_ytd_chart = ytd_naechte[[land_auswahl]].reset_index().rename(columns={land_auswahl:'Nächte YTD'})
    df_ytd_chart['Monat'] = df_ytd_chart['Monat'].dt.strftime('%B %Y')
    df_ytd_chart = df_ytd_chart.sort_values('Monat')
    chart_ytd = alt.Chart(df_ytd_chart).mark_bar(color='orange').encode(
        x=alt.X('Monat:N', sort=list(df_ytd_chart['Monat'])),
        y=alt.Y('Nächte YTD:Q'),
        tooltip=['Monat','Nächte YTD']
    ).properties(width=700)
    st.altair_chart(chart_ytd)






# streamlit run guest.py

# Aus den eingegebenen Werten sollten folgende Zahlen ausgerechnet werden und in einer immer sichtbaren Tabelle dargestellt werden: 1. Durchschnittliche Anzahl Nächte pro Anreise pro Land (Nächte / Anreisen) 2. % Anteil Nächte pro Land (Total alle Länder = 100%) 3. % Veränderung der Anzahl Nächte im Vergleich zum selben Monat im Vorjahr pro Land 4. alle oben aufgeführten Berechnungen sollten auch als YTD berechnet werden (z.B. März beinhaltet die Daten von Januar, Februar und März; Mai beinhaltet die Daten von Januar, Februar, März, April und Mai) Zudem sollten auch Einträge mit dem Wert 0 akzeptiert werden.

# Eingabeformular
#st.header("Neue Daten eingeben")
#with st.form("eingabe_form"):
    #jahr = st.selectbox("Jahr", list(range(2017, datetime.today().year + 4)), index=list(range(2017, datetime.today().year + 4)).index(datetime.today().year))
    #monat = st.selectbox("Monat", list(range(1, 13)), index=datetime.today().month - 2)
